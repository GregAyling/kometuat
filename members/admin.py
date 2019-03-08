from django.contrib import admin
from .models import Member
from datetime import date
from django.http import HttpResponse

COLUMN_NAMES = [
            "last_name",
            "date_of_birth",
            "gender",
            "section",
            "address_1",
            "address_2",
            "adult_investiture",
            "medical_alert",
            "carer_alert",
            "financial_alert",
            "general_alert",
            "alert_notes",
            "allergies",
            "ambulance_fund",
            "app_form_issued",
            "app_form_returned",
            "app_sent_to_branch",
            "appointment_review",
            "blue_card_wwcc_first_issued",
            "blue_card_wwcc_next_expires",
            "blue_card_wwcc_number",
            "carer_1_blue_card_expiry",
            "carer_1_blue_card_no",
            "carer_1_email",
            "carer_1_employer",
            "carer_1_fax",
            "carer_1_help_offer",
            "carer_1_interests",
            "carer_1_mobile",
            "carer_1_name",
            "carer_1_occupation",
            "carer_1_previous_scouting",
            "carer_1_relationship",
            "carer_1_skills",
            "carer_1_work_phone",
            "carer_2_blue_card_exp",
            "carer_2_blue_card_no",
            "carer_2_email",
            "carer_2_employer",
            "carer_2_fax",
            "carer_2_help_offer",
            "carer_2_interests",
            "carer_2_mobile",
            "carer_2_name",
            "carer_2_occupation",
            "carer_2_previous_scouting",
            "carer_2_relationship",
            "carer_2_skills",
            "carer_2_work_phone",
            "change_type",
            "contact_phone",
            "contact_phone1",
            "cpr_date",
            "cub_investiture",
            "date_left_scouting",
            "dietary_preferences",
            "disabilities",
            "doctors_phone",
            "email_type",
            "emergency_contact",
            "employer_school",
            "expiry_date",
            "family_doctor",
            "family_notes",
            "first_aid_cert",
            "first_invested",
            "first_names",
            "heard_about_scouts",
            "home_fax",
            "home_language",
            "home_phone",
            "image_file",
            "interests",
            "joey_investiture",
            "last_review_date",
            "last_tetanus_date",
            "main_email_address",
            "medical_action_plan",
            "medical_conditions",
            "medical_fund",
            "medicare_number",
            "medications",
            "member_type",
            "membership_number",
            "mobile",
            "nationality",
            "occupation",
            "place_of_birth",
            "position_on_card",
            "post_address_1",
            "post_address_2",
            "postal_postcode",
            "postal_state",
            "postal_suburb",
            "postcode",
            "pref_strength",
            "preferred_name",
            "qld_branch_family_code",
            "quick_initials",
            "rank",
            "record_created",
            "relationship_to_member",
            "religion",
            "residential_situation",
            "rover_investiture",
            "salutation",
            "scout_investiture",
            "scout_name",
            "sighted_blue_card",
            "six_patrol_group",
            "social_media",
            "source",
            "state",
            "suburb",
            "venturer_investiture",
            "waiting_list_comment",
            "waiting_list_date",
            "where_to",
            "work_fax",
            "work_phone",
            "yh_appointment",
]

def column_value(object_name):
        from django.utils.encoding import smart_str
        X = smart_str(object_name)
        if X == "None":
                return ""
        else:
                return X 

def export_csv(modeladmin, request, queryset):
        import csv
        from django.utils.encoding import smart_str
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=members.csv'
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow(list(map(lambda X: X.upper(),COLUMN_NAMES))) # Output column headers as uppercase column names.
        for obj in queryset:
                writer.writerow(list(map(lambda X: column_value(getattr(obj,X)),COLUMN_NAMES)))
        return response

def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=members.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("Members")
    
    row_num = 0
    
    columns = list(map(lambda X:(X.upper(), 10000),COLUMN_NAMES)) # Output column headers as column names with fixed width.

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = list(map(lambda X: getattr(obj,X),COLUMN_NAMES))
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response
    
def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Members"

    row_num = 0

    columns = list(map(lambda X:X.upper(),COLUMN_NAMES)) # Output column headers as uppercase column names.

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num]
        c.font = Font(bold=True)
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = 30

    for obj in queryset:
        row_num += 1
        row = list(map(lambda X: getattr(obj,X),COLUMN_NAMES))
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.alignment = c.alignment.copy(wrap_text=True)

    wb.save(response)
    return response

def list_email_addresses(modeladmin, request, queryset):
        # Add all individual email addresses to a set so that duplicates are eliminated.
        # The "main_email_address" field may contain multiple addresses separated by a semi-colon.
        email_address_set = set()
        for obj in queryset:
                obj_email_list = column_value(obj.main_email_address).split(";")
                for email_address in obj_email_list:
                        email_address_set.add(email_address)
        # Convert the set to a string of email addresses separated by a semi-colon.
        email_address_string = '; '.join(list(sorted(email_address_set)))
        # Display email address list as a user message.
        modeladmin.message_user(request, "Email Address List: %s" % email_address_string)
        return

class MemberAdmin(admin.ModelAdmin):
        
        actions = [
                list_email_addresses, 
                export_csv, 
                export_xls, 
                export_xlsx, ]
        fieldsets = [
        ('Personal Details', {'fields': [ 
                'last_name', 
                'first_names', 
                'date_of_birth', 
                'place_of_birth',
                'gender', 
                'preferred_name', 
                'religion', 
                'nationality', 
                'interests', 
                'occupation', 
                'employer_school', 
                'home_language',
        ], 'classes': ['collapse']}),
        ('Address', {'fields': [ 
                'address_1', 
                'address_2', 
                'suburb',
                'state', 
                'postcode',
        ], 'classes': ['collapse']}),
        ('Postal Address', {'fields': [  
                'post_address_1', 
                'post_address_2', 
                'postal_suburb',
                'postal_state', 
                'postal_postcode',
        ], 'classes': ['collapse']}),
        ('Contact Information', {'fields': [ 
                'main_email_address', 
                'home_phone', 
                'home_fax',  
                'work_phone', 
                'work_fax', 
                'mobile',  
                'email_type', 
                'heard_about_scouts', 
                'source', 
                'quick_initials', 
                'image_file', 
                'section', 
                'six_patrol_group', 
                'member_type', 
                'scout_name', 
                'rank', 
                'social_media',
        ], 'classes': ['collapse']}),
        ('Progression', {'fields': [ 
                'record_created', 
                'waiting_list_comment', 
                'waiting_list_date', 
                'app_form_issued', 
                'app_form_returned', 
                'first_invested',
                'joey_investiture', 
                'cub_investiture',
                'scout_investiture', 
                'venturer_investiture', 
                'rover_investiture', 
                'yh_appointment', 
                'adult_investiture', 
                'app_sent_to_branch',
                'membership_number',
                'date_left_scouting', 
                'where_to',
        ], 'classes': ['collapse']}),
        ('Adult Service', {'fields': [   
                'appointment_review',
                'last_review_date', 
                'sighted_blue_card', 
                'first_aid_cert',  
                'cpr_date', 
                'blue_card_wwcc_number', 
                'blue_card_wwcc_first_issued', 
                'blue_card_wwcc_next_expires',
        ], 'classes': ['collapse']}),
        ('Family Details', {'fields': [ 
                'salutation', 
                'residential_situation', 
                'family_notes', 
                'qld_branch_family_code',
        ], 'classes': ['collapse']}),
        ('Member Alerts', {'fields': [  
                'medical_alert', 
                'carer_alert', 
                'financial_alert', 
                'general_alert', 
                'alert_notes',
        ], 'classes': ['collapse']}),
        ('Carer 1 / Spouse', {'fields': [ 
                'carer_1_name', 
                'carer_1_relationship', 
                'carer_1_occupation',
                'carer_1_employer', 
                'carer_1_work_phone', 
                'carer_1_mobile',  
                'carer_1_fax', 
                'carer_1_email', 
                'carer_1_blue_card_no',
                'carer_1_blue_card_expiry', 
                'carer_1_skills',  
                'carer_1_interests', 
                'carer_1_help_offer', 
                'carer_1_previous_scouting',
        ], 'classes': ['collapse']}),
        ('Carer 2', {'fields': [ 
                'carer_2_name', 
                'carer_2_relationship', 
                'carer_2_occupation',
                'carer_2_employer', 
                'carer_2_work_phone', 
                'carer_2_mobile',  
                'carer_2_fax', 
                'carer_2_email', 
                'carer_2_blue_card_no',
                'carer_2_blue_card_exp', 
                'carer_2_skills',  
                'carer_2_interests', 
                'carer_2_help_offer', 
                'carer_2_previous_scouting',
        ], 'classes': ['collapse']}),
        ('Medical Details', {'fields': [  
                'medicare_number',  
                'expiry_date', 
                'position_on_card',
                'medical_fund', 
                'ambulance_fund',  
                'family_doctor', 
                'doctors_phone',
                'last_tetanus_date',
        ], 'classes': ['collapse']}),
        ('Known Conditions', {'fields': [ 
                'allergies', 
                'medical_conditions', 
                'disabilities',  
                'medications',  
                'medical_action_plan',
        ], 'classes': ['collapse']}),
        ('Second Emergency Contact', {'fields': [ 
                'emergency_contact', 
                'relationship_to_member', 
                'contact_phone',
        ], 'classes': ['collapse']}),
        ('Dietary Notes', {'fields': [ 
                'dietary_preferences', 
                'pref_strength',
        ], 'classes': ['collapse']}),
        ]
        def full_name(self,obj):
                return (obj.first_names if obj.first_names is not None else "{No first names}") + " " + (obj.last_name if obj.last_name is not None else "{No last name}")
        def age(self,obj):
                if obj.date_of_birth == None:
                        return None
                today = date.today()
                birth_date = obj.date_of_birth
                today = date.today()
                age = round((today - birth_date).days/365.25,1)
                return age
        list_display = (
                'full_name', 
                'date_of_birth', 
                'age', 
                'section', 
                'six_patrol_group', 
                'medical_alert', 
                'carer_alert', 
                'financial_alert', 
                'general_alert',
                )
        list_filter = (
                'section', 
                'medical_alert', 
                'carer_alert', 
                'financial_alert', 
                'general_alert',
                )
        ordering = (
                'first_names', 
                'last_name',
                )
        search_fields = [
                'last_name', 
                'first_names',
                ]
        list_per_page = 15

admin.site.register(Member, MemberAdmin)
